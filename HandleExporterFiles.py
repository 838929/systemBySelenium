#ABOUT : this module copy data from 'parts' file and paste to other excel  files list.

import openpyxl

files_path = r'C:\Users\ahmed.mosaad\Desktop\Rs\system\exporter_files\\'

def get_parts_comps(xlfile = "parts", sheet_name = "Sheet1", parts_column = 1, com_column = 2):
	# initialize list for urls ".xml"
	parts_list = []
	com_list = []
	wb = openpyxl.load_workbook(files_path + xlfile + ".xlsx")
	active_sheet = wb.get_sheet_by_name(sheet_name)
	rows_num = active_sheet.get_highest_row()

	for i in range(2, rows_num+1):
		parts_list.append(active_sheet.cell(row = i, column = parts_column).value)
		com_list.append(active_sheet.cell(row = i, column = com_column).value)
	return parts_list, com_list


def drop_parts_comps(parts_list, companies_list, xlfile, sheet_name , parts_column_no, companies_column_no ):
	wb = openpyxl.load_workbook(xlfile + ".xlsx")
	active_sheet = wb.get_sheet_by_name(sheet_name)
	rows_num = active_sheet.get_highest_row()

	# delete old data .
	k = 0
	for j in range(2, rows_num+2):
		active_sheet.cell(row = j, column = parts_column_no).value = None
		active_sheet.cell(row = j, column = companies_column_no).value = None
		k += 1

	# insert new data in cells .
	k = 0
	for j in range(2, len(parts_list)+2):
		active_sheet.cell(row = j, column = parts_column_no).value = parts_list[k]
		active_sheet.cell(row = j, column = companies_column_no).value = companies_list[k]
		k += 1

	wb.save(xlfile + ".xlsx")

files = ['Pcn', 'OBS', 'Notmatch', 'Mask']

def main():
	PNs, COMs = get_parts_comps()
	for file in files:
		drop_parts_comps(PNs, COMs, files_path + file, 'Sheet1', 1, 2)
	print('PNs, COMs copied to files')

# run script 
# main()